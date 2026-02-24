from datetime import datetime
from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


# ================= CONSTANTS =================

ALL_TIME_SLOTS = [
    "5 AM - 6 AM",
    "6 AM - 7 AM",
    "7 AM - 8 AM",
    "8 AM - 9 AM",
    "9 AM - 10 AM",
    "10 AM - 11 AM",
    "11 AM - 12 PM",
    "12 PM - 1 PM",
    "1 PM - 2 PM",
    "2 PM - 3 PM",
    "3 PM - 4 PM",
    "4 PM - 5 PM",
    "5 PM - 6 PM",
    "6 PM - 7 PM",
    "7 PM - 8 PM",
    "8 PM - 9 PM",
    "9 PM - 10 PM",
]


# ================= HELPERS =================

def normalize_slot(slot: str) -> str:
    """
    Normalizes time slot strings so formats like:
    '4:00 PM - 5:00 PM' and '4 PM - 5 PM'
    become identical keys.
    """
    if not slot:
        return ""

    slot = slot.replace(":00", "")
    slot = slot.replace("AM", " AM").replace("PM", " PM")
    slot = " ".join(slot.split())
    return slot.strip()


def auto_adjust_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def auto_adjust_row_heights(ws, base_height=18):
    """
    Adjust row height based on wrapped text.
    Uses line-break count as heuristic.
    """
    for row in ws.iter_rows():
        max_lines = 1
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                lines = cell.value.count("\n") + 1
                max_lines = max(max_lines, lines)
        ws.row_dimensions[row[0].row].height = max_lines * base_height


# ================= WEEKLY ACTIVITY SHEET =================

def build_weekly_activity_sheet(ws, activities):
    activity_map = defaultdict(dict)
    dates = set()

    # Build lookup map
    for a in activities:
        date = a.get("date")
        slot = normalize_slot(a.get("slot", ""))
        chat = a.get("chat", "")
        activity_map[date][slot] = chat
        dates.add(date)

    if not dates:
        return

    # Sort dates chronologically
    sorted_dates = sorted(
        list(dates),
        key=lambda d: datetime.strptime(d, "%Y-%m-%d")
    )

    # Split into weeks (7 days per week)
    weeks = [
        sorted_dates[i:i + 7]
        for i in range(0, len(sorted_dates), 7)
    ]

    row = ws.max_row + 1

    for week_index, week_dates in enumerate(weeks, start=1):
        # Week title
        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=len(week_dates) + 1
        )
        title_cell = ws.cell(row=row, column=1)
        title_cell.value = f"Week {week_index}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        row += 2

        # Header row
        ws.cell(row=row, column=1).value = "Time Slot"
        ws.cell(row=row, column=1).font = Font(bold=True)

        for col, date_str in enumerate(week_dates, start=2):
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            header = dt.strftime("%d %b. %Y\n(%A)")
            c = ws.cell(row=row, column=col)
            c.value = header
            c.font = Font(bold=True)
            c.alignment = Alignment(
                wrap_text=True,
                horizontal="center",
                vertical="center"
            )

        row += 1

        # Time slots
        for raw_slot in ALL_TIME_SLOTS:
            slot_key = normalize_slot(raw_slot)
            ws.cell(row=row, column=1).value = raw_slot

            for col, date_str in enumerate(week_dates, start=2):
                ws.cell(
                    row=row,
                    column=col
                ).value = activity_map[date_str].get(slot_key, "")

            row += 1

        row += 2

    auto_adjust_columns(ws)
    auto_adjust_row_heights(ws)


# ================= THOUGHT RECORDS SHEET =================

def build_thought_records_sheet(ws, thoughts):
    headers = [
        "Date & Time",
        "Trigger",
        "Feeling",
        "Negative Thought",
        "New Thought",
        "Outcome",
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            wrap_text=True,
            vertical="center",
            horizontal="center"
        )

    for t in thoughts:
        raw_ts = t.get("created_at", "")
        formatted_dt = ""

        if raw_ts:
            try:
                dt = datetime.fromisoformat(raw_ts)
                formatted_dt = dt.strftime("%d %b. %Y\n%I:%M %p")
            except Exception:
                formatted_dt = raw_ts

        ws.append([
            formatted_dt,
            t.get("trigger", ""),
            t.get("feeling", ""),
            t.get("negative_thought", ""),
            t.get("new_thought", ""),
            t.get("outcome", ""),
        ])

        ws.cell(
            row=ws.max_row,
            column=1
        ).alignment = Alignment(
            wrap_text=True,
            vertical="center"
        )

    auto_adjust_columns(ws)
    auto_adjust_row_heights(ws)


# ================= MAIN ENTRY =================

def generate_excel_report(data: dict) -> bytes:
    wb = Workbook()

    # ---------- Weekly Activity ----------
    ws1 = wb.active
    ws1.title = "Weekly Activity"

    ws1.append(["Patient Name", data["patient"]["name"]])
    ws1.append(["Email", data["patient"]["email"]])
    ws1.append([])

    if data.get("include_weekly_activity"):
        build_weekly_activity_sheet(
            ws1,
            data.get("weekly_activities", [])
        )

    # ---------- Thought Records ----------
    if data.get("include_thought_records"):
        ws2 = wb.create_sheet("Thought Records")
        build_thought_records_sheet(
            ws2,
            data.get("thought_records", [])
        )

    # ---------- Save ----------
    buffer = BytesIO()
    wb.save(buffer)
    wb.close()
    return buffer.getvalue()
